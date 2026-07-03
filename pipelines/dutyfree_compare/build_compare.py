#!/usr/bin/env python3
"""면세점 100ml당 가격 비교: 신라 vs 롯데 vs 신세계(SSG) → 엑셀(.xlsx) + CSV (CMPA-650).

보드 요청(부모 CMPA-649): 신라면세·롯데면세·신세계면세 가격을 각각 100ml당 가격(USD)으로
환산해 세 면세점을 같은 제품(직접 이름 매칭) 단위로 비교하고 엑셀 테이블로 저장한다.

방법론(이슈 CMPA-650 준수):
- 용량 파싱: 신라는 `scripts/normalize_dataset.extract_volume_ml(위스키명)` 재사용,
  롯데/SSG는 CSV `volume_ml` 그대로. 용량 못 구하면 그 행은 비교 제외(임의 700ml 금지).
- 100ml당 가격 = 가격 / volume_ml * 100. 기준 통화 = USD(면세 USD 표기, FX 노이즈 없음).
  가격 선택: 롯데/SSG=sale_price, 신라=표시가_USD (analyze_attractiveness 일관). 할인가는 보조 병기.
- 매칭: 다단계 직접 이름 매처 (canonical_id 조인 대신)
  Stage1: norm() 정규화 키 완전 일치
  Stage2: 제품명 변형(RB코드·영문표기 통일 등) 적용 후 재매칭
  Stage3: 하드코딩된 ALIAS_MAP 적용
  CMPA-177: 년수/CS/피티드/셰리/버번 에디션 토큰 다르면 절대 병합 금지
- 노이즈: 잔세트/번들은 `pipelines.common.whisky_quality.is_bundle_noise`로 제외.
- 비교 시트 포함 기준: 3개 면세점 중 2개 이상 매칭(at least 2-of-3).
- 국내최저가(보드 추가 요청 2026-06-28): **데일리샷 + 트레이더스 + 코스트코** 합산 최저가.
  normalized_prices.csv 에서 소스별 최신가 중 최소(per_source_latest_floor 재사용, CMPA-496).
  canonical_id 로 비교행에 조인하고, 단일 환율로 면세 최저 100ml당가(₩환산)와의 차이를 표기.
  (국내 표준용량=canonical volume — analyze_attractiveness 면세매력도 방법론과 동일. 1L 변형은
   per-100ml 에 약간의 용량 가정 오차가 있을 수 있어 메타에 단서.)

산출물(멱등·재실행 가능):
- reports/dutyfree-compare/면세점_100ml_가격비교_<date>.xlsx
  (시트: 비교/신라단독/롯데단독/신세계단독/메타)
- reports/dutyfree-compare/면세점_100ml_가격비교_<date>.csv   (UTF-8 BOM, 비교 시트 미러)

주의: 면세가는 출국/면세 한도(2병·2L·$400) 조건. '수집일 기준'값이며 현재가 단정 아님.
"""
import argparse
import csv
import datetime
import os
import re
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", ".."))

# 정본 헬퍼 재사용(중복구현 금지)
sys.path.insert(0, os.path.join(ROOT, "scripts"))
sys.path.insert(0, os.path.join(ROOT, "pipelines", "shilla_dutyfree"))
from normalize_dataset import extract_volume_ml          # noqa: E402
from analyze_attractiveness import norm, load_canonical, load_fx  # noqa: E402
from pipelines.common.whisky_quality import is_bundle_noise  # noqa: E402
from pipelines.common.source_floor import per_source_latest_floor  # noqa: E402

NORMALIZED_CSV = os.path.join(ROOT, "data", "whisky-prices", "normalized",
                              "normalized_prices.csv")
# 보드 지정(CMPA-650 2026-06-28): 국내최저가 = 데일리샷 + 트레이더스 + 코스트코 합산 최저.
# (이마트/롯데마트/기타 마트는 제외 — 보드가 명시한 3개 소스만.)
DOMESTIC_SOURCES = {"데일리샷", "트레이더스", "코스트코"}

STALE_WARN_DAYS = 3  # 최신 스냅샷이 이 일수 이상 오래되면 경고(데이터 3원칙: 수집날짜 신뢰 신호)


def _kst_today():
    return (datetime.datetime.now(datetime.timezone.utc)
            + datetime.timedelta(hours=9)).date()


def _date_from_path(path: str) -> str:
    """스냅샷 파일명에서 YYYY-MM-DD 날짜 토큰을 추출. (신라는 접미사, 마트는 접두사 모두 매칭)."""
    m = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(path or ""))
    return m.group(1) if m else ""


def _latest_snapshot(pattern: str, required: bool = True):
    """glob pattern 에서 파일명 날짜순(YYYY-MM-DD) 최신 파일 경로를 반환.

    하드코딩 날짜 대신 매일 자동으로 최신 스냅샷을 집어 일간 루틴이 stale 되지 않게 한다.
    (CMPA-672) 파일명에서 날짜를 추출해 정렬하므로 접두/접미 위치와 무관하다.
    required=False 면 없을 때 ("", "") 를 반환(예: 데일리샷 보강 캐시 — 없어도 비치명).
    """
    import glob as _glob
    files = [f for f in _glob.glob(pattern) if _date_from_path(f)]
    if not files:
        if required:
            raise FileNotFoundError(f"스냅샷 없음: {pattern}")
        return "", ""
    path = max(files, key=_date_from_path)
    return path, _date_from_path(path)


def _staleness_warn(date_str: str, label: str, today=None,
                    max_age_days: int = STALE_WARN_DAYS) -> bool:
    """최신 스냅샷이 max_age_days 일 이상 오래되면 stderr 경고. 경고했으면 True.

    today 를 주입할 수 있어 테스트가 결정론적이다(기본=KST 오늘)."""
    if not date_str:
        return False
    try:
        d = datetime.date.fromisoformat(date_str)
    except (ValueError, TypeError):
        return False
    today = today or _kst_today()
    age = (today - d).days
    if age >= max_age_days:
        print(f"⚠️ STALE: {label} 최신 스냅샷이 {date_str}({age}일 전)입니다 — "
              f"수집 루틴 점검 필요. 현재가 단정 금지(데이터 3원칙).", file=sys.stderr)
        return True
    return False


LOTTE_CSV, LOTTE_DATE = _latest_snapshot(
    os.path.join(ROOT, "assets", "lotte_dutyfree", "snapshots",
                 "????-??-??_lotte_whisky.csv"))
SHILLA_CSV, SHILLA_DATE = _latest_snapshot(
    os.path.join(ROOT, "data", "shilla-dutyfree", "신라면세_위스키_????-??-??.csv"))
SSG_CSV, SSG_DATE = _latest_snapshot(
    os.path.join(ROOT, "assets", "ssg_dutyfree", "snapshots",
                 "????-??-??_ssg_whisky.csv"))


def warn_all_staleness(today=None):
    """세 면세점 + 데일리샷 보강 스냅샷의 staleness 를 한 번에 점검(경고만)."""
    n = 0
    for d, label in ((SHILLA_DATE, "신라면세"), (LOTTE_DATE, "롯데면세"),
                     (SSG_DATE, "신세계(SSG)면세"), (ONLINE_DATE, "데일리샷 온라인 보강")):
        if _staleness_warn(d, label, today=today):
            n += 1
    return n


# ── Stage2/Stage3 규칙 ────────────────────────────────────────────────────────
# CMPA-177 보호 토큰: 이 값이 서로 다른 두 이름은 절대 같은 제품이 아님
# 정규식 패턴으로 이름 안의 년수/CS/피티드 토큰을 추출
_AGE_PAT = re.compile(r"(\d+)[yY년]")
_CS_PAT = re.compile(r"\b(cs|캐스크스트렝스|캐스크스트랭스|caskstrength)\b")
_PEATED_PAT = re.compile(r"(피티드|peated|peat)")
_SHERRY_PAT = re.compile(r"(셰리|sherry)")
_BOURBON_PAT = re.compile(r"(버번|bourbon)")


def _age_token(key: str) -> int | None:
    """norm() 적용 후 key에서 숫자년/Y 토큰 추출. 없으면 None."""
    m = _AGE_PAT.search(key)
    return int(m.group(1)) if m else None


def _flag(key: str, pat: re.Pattern) -> bool:
    return bool(pat.search(key))


def _cmpa177_ok(key_a: str, key_b: str) -> bool:
    """두 키가 CMPA-177 기준으로 병합 가능하면 True.
    - 년수: 둘 다 존재하면 동일해야 함 (하나만 없으면 통과 — 대표 표기 차이)
    - CS/피티드/셰리/버번: 한쪽에만 있으면 병합 금지
    - 프루프 익스프레션(101/107/114 등): 집합이 다르면 다른 제품(아래 _proof_tokens 참고)
    """
    age_a = _age_token(key_a)
    age_b = _age_token(key_b)
    if age_a is not None and age_b is not None and age_a != age_b:
        return False
    # 에디션 토큰 비대칭 가드
    for pat in (_CS_PAT, _PEATED_PAT, _SHERRY_PAT, _BOURBON_PAT):
        if _flag(key_a, pat) != _flag(key_b, pat):
            return False
    return True


# 모델/프루프 익스프레션 숫자(메이커스 마크 46/101, 와일드터키 101, 버팔로트레이스 107,
# 글렌파클라스 105 등). 같은 브랜드의 서로 다른 익스프레션을 가르는 결정 토큰이다.
# 40~129 범위의 독립 2~3자리 정수만 본다 — 나이(10~30)·에디션 서수(≤10)·용량(700/750/1000)·
# 빈티지 연도(4자리)·도수(%·'도')는 제외한다.
# ⚠️ norm() 후엔 '101'+'1000ml' 가 '1011000ml' 로 붙어 파싱 불가 → 반드시 공백이 남은
#    원문(raw) 이름에서 단어경계로 추출한다(_model_nums).
_MODEL_NUM_PAT = re.compile(r"(?<!\d)(\d{2,3})(?!\d)")


def _model_nums(raw_name: str) -> frozenset:
    """원문(공백 보존) 이름에서 모델/프루프 결정 숫자 집합(40~129)을 추출.
    직후가 ml/l(용량)·%/도(도수)인 숫자는 제외. 예: '메이커스 마크 46'→{46},
    '메이커스 마크 101 1000ml'→{101}, '발렌타인 21년'→{}(나이), '아티스트 에디션 5'→{}(서수),
    '글렌파클라스 105'→{105}, '듀어스 12년 43%'→{}."""
    out = set()
    for m in _MODEL_NUM_PAT.finditer(raw_name or ""):
        n = int(m.group(1))
        if not (40 <= n <= 129):
            continue
        tail = (raw_name[m.end():m.end() + 3] or "").lower().lstrip()
        if tail.startswith("ml") or tail[:1] == "l":
            continue  # 용량(예: 100ml / 100L)
        if tail[:1] in ("%", "도") or tail.startswith("proof"):
            continue  # 도수(예: 46% / 50도)
        out.add(n)
    return frozenset(out)


def _proof_ok(raw_a: str, raw_b: str) -> bool:
    """두 원문 이름이 같은 모델/프루프 익스프레션이면 True(집합 동일). CMPA-669 오매칭 가드."""
    return _model_nums(raw_a) == _model_nums(raw_b)


def _vol_compatible(vol_a: int, vol_b: int) -> bool:
    """용량 호환: 1000ml ≈ 1L 이면 같은 제품. 동일하거나 1000↔1000 변환."""
    def _norm_vol(v):
        return 1000 if v == 1 else v
    return _norm_vol(vol_a) == _norm_vol(vol_b)


# ── Stage2 변환 함수 ────────────────────────────────────────────────────────────
# norm() 적용 후 key에 추가로 적용하는 정규화. 결과 키를 stage2_key로 쓴다.
_RB_PAT = re.compile(r"\d{2}rb")              # (24RB), (23RB) 등 출시 코드
_VAP_PAT = re.compile(r"vap$")                # 끝의 VAP 마케팅 코드
_FESTIVE_PAT = re.compile(r"festive")         # festive 에디션 표기
_SLIM_PAT = re.compile(r"슬림팩?")              # 슬림팩 유통 포장
_CNY_PAT = re.compile(r"cny\d*")             # CNY 설 에디션 한정 포장
_IGL_PAT = re.compile(r"igl")                # IGL 인증 표기
_NEW_EDITION_PAT = re.compile(r"(newedition|뉴에디션|뉴에디션)")


def _stage2_key(key: str) -> str:
    """norm() 결과 키에 Stage2 변환을 적용해 더 표준화된 비교 키를 반환."""
    k = key
    k = _RB_PAT.sub("", k)       # 24rb, 23rb 등 제거
    k = _VAP_PAT.sub("", k)      # 끝의 vap 제거
    k = _FESTIVE_PAT.sub("", k)  # festive 제거
    k = _SLIM_PAT.sub("", k)     # 슬림팩 제거
    k = _CNY_PAT.sub("", k)      # cny2026 등 제거
    k = _IGL_PAT.sub("", k)      # igl 제거

    # 블루 → 블루라벨 (단, 이미 블루라벨이면 변환 불필요)
    k = re.sub(r"블루(?!라벨)", "블루라벨", k)

    # 더 글렌리벳 ↔ 글렌리벳 ("더" 접두사 제거)
    k = re.sub(r"^더글렌리벳", "글렌리벳", k)

    # 더 글렌그란트 ↔ 글렌그란트 ("더" 접두사 제거)
    k = re.sub(r"^더글렌그란트", "글렌그란트", k)

    # 더 글렌그런트 등 표기 변형
    k = re.sub(r"^더글렌", "글렌", k)

    # cask → 캐스크 음역 통일 (영문 cask만 한글로)
    k = re.sub(r"cask(?!strength)", "캐스크", k)

    # PX Cask / PX캐스크 통일
    k = re.sub(r"px캐스크", "px캐스크", k)    # 이미 같음, no-op
    k = re.sub(r"pxcask", "px캐스크", k)

    # grand cru / 그랑 크루 / 그랑크뤼 → grandcru 통일
    k = re.sub(r"그랑크뤼", "그랑크루", k)
    k = re.sub(r"grandcru", "그랑크루", k)

    # Grand Couronne / 그랑 쿠론 → 그랑쿠론 (이미 norm에 의해 공백 제거됨)
    k = re.sub(r"gc(?=700ml|1000ml|$)", "그랑쿠론", k)  # GC = Grand Couronne 축약

    # Vat3 / vat3 → 벳3 통일
    k = re.sub(r"vat(\d)", r"벳\1", k)

    # Perpetual / perpetual → 퍼페추얼
    k = re.sub(r"perpetual", "퍼페추얼", k)

    # timeseries / 타임시리즈 → 타임시리즈
    k = re.sub(r"timeseries", "타임시리즈", k)

    # 그랑레제르바 / 그랑리제르바 — 동의어 ("gran reserva")
    k = re.sub(r"그랑리제르바", "그랑레제르바", k)
    k = re.sub(r"granreserva", "그랑레제르바", k)

    # 코르테즈 / 코르테스 철자 변형 → 코르테스 (Grand Cortes)
    k = re.sub(r"코르테즈", "코르테스", k)

    # 루미나리 / 루미너리 (Luminary) 철자 변형
    k = re.sub(r"루미나리", "루미너리", k)

    # 포트폴리오 시리즈2 / 시리즈 → 시리즈 (숫자 배치 코드 제거)
    k = re.sub(r"시리즈\d+", "시리즈", k)

    # newedition / 뉴에디션 / 뉴 에디션 → 뉴에디션
    k = re.sub(r"newedition", "뉴에디션", k)

    # 105cs ↔ 105 (Glenfarclas 105 = 105 CS 동일 제품)
    k = re.sub(r"105cs", "105", k)

    # 달모어 시가 몰트 / 시가몰트 → 시가몰트
    k = re.sub(r"시가몰트", "시가몰트", k)  # no-op
    k = re.sub(r"시가몰트", "시가몰트", k)

    # 달모어 45년 2024년 에디션 → 달모어 45년 (연도 에디션 제거)
    k = re.sub(r"(\d+년)20\d{2}년에디션", r"\1", k)

    # 1L → 1000ml 통일
    k = re.sub(r"1l(?=$)", "1000ml", k)

    # 용량 앞 숫자+ml 형태 재검토: "2007" 같은 빈티지 연도는 제거 대상 아님

    return k


# ── Stage3 하드코딩 별칭 맵 ────────────────────────────────────────────────────
# CMPA-177 안전 확인 후 수동으로 검증된 동일 제품 쌍.
# 키: stage2_key(norm(이름A)) → 값: stage2_key(norm(이름B))
# 방향: 소수 표기에서 다수 표기로 합치기.
# 검증 출처: 웹 검색 2026-06-28
ALIAS_MAP: dict[str, str] = {
    # 조니워커 블루 = 블루라벨 (Stage2 블루→블루라벨 처리로 이미 해결됨)
    # 여기서는 Stage2 후에도 남는 차이를 처리

    # 글렌피딕 21년 (신라 단순표기) = 글렌피딕 21년 그랑 레제르바
    # stage2 후: 글렌피딕21년700ml ↔ 글렌피딕21년그랑레제르바700ml
    # → 신라가 그랑레제르바 이름 생략한 것. 동일 제품(확인됨).
    "글렌피딕21년700ml": "글렌피딕21년그랑레제르바700ml",

    # 글렌피딕 26년 GC = 그랑 쿠론 (Stage2 gc→그랑쿠론 이미 처리)
    # stage2 후: 글렌피딕26년그랑쿠론700ml ↔ 글렌피딕26그랑쿠론700ml (년 차이)
    "글렌피딕26그랑쿠론700ml": "글렌피딕26년그랑쿠론700ml",

    # 달모어 시가 몰트 ↔ 달모어 시가몰트 (공백 차이는 norm에서 제거됨)
    # 그러나 용량이 다를 수 있음 (1000ml vs 700ml) → 용량 체크에서 필터됨

    # 달모어 루미너리 No.2 ↔ 루미나리 (Stage2에서 이미 처리)

    # 글렌피딕 23년 그랑크뤼 (stage2→그랑크루) = 글렌피딕 23년 그랑크루
    # 슬림팩은 제거됨 → 동일 키
    # 이미 Stage2에서 처리됨

    # 더 글렌리벳 트리플 캐스크 머추어드 ↔ 화이트 오크 버전
    # 화이트 오크는 다른 마감재 — CMPA-177에 의해 다른 제품으로 처리

    # 달모어 포트폴리오 시리즈 2007년 ↔ 포트폴리오 시리즈2 2007
    # stage2 후: 달모어포트폴리오시리즈2007년700ml ↔ 달모어포트폴리오시리즈2007700ml
    "달모어포트폴리오시리즈2007년700ml": "달모어포트폴리오시리즈2007700ml",

    # 글렌알라키 8년 스코티시 오크 피니쉬 싱글몰트 ↔ 글렌알라키 8년
    # 이건 SSG가 상세 표기, Lotte는 단순 표기 — 동일 제품
    # 단, 피니쉬 캐스크가 다른 버전일 수 있어 신중히 처리
    # norm 후 비교: 글렌알라키8년스코티시오크피니쉬싱글몰트700ml vs 글렌알라키8년700ml
    # 스코티시 오크 피니쉬 = 공식 제품명의 일부 (GlenAllachie 8yo Scottish Oak Finish)
    # 롯데에서는 단순히 "글렌알라키 8년" — 같은 제품
    "글렌알라키8년스코티시오크피니쉬싱글몰트700ml": "글렌알라키8년700ml",

    # 글렌피딕 30년 타임시리즈 (stage2 처리됨) vs 글렌피딕 30 타임시리즈 (년 없음)
    "글렌피딕30타임시리즈700ml": "글렌피딕30년타임시리즈700ml",

    # 달모어 45년 2024년 에디션 (SSG) → 달모어 45년 (신라 단순 표기)
    # Stage2에서 "20XX년에디션" 제거 → 달모어45년700ml 동일
    # (이미 Stage2에서 처리됨)

    # 달모어 루미너리 (SSG: 루미나리) — Stage2에서 이미 처리
}


def _apply_alias(key: str) -> str:
    """ALIAS_MAP 양방향 조회."""
    if key in ALIAS_MAP:
        return ALIAS_MAP[key]
    # 역방향도 확인
    for k, v in ALIAS_MAP.items():
        if v == key:
            return k
    return key


def _make_key(raw_name: str) -> tuple[str, str]:
    """(stage1_key, stage2_key) 반환."""
    k1 = norm(raw_name)
    k2 = _stage2_key(k1)
    k2 = _apply_alias(k2)
    return k1, k2


def _f(v):
    try:
        x = float(v)
        return x if x > 0 else None
    except (ValueError, TypeError):
        return None


# ── 데이터 로더 ────────────────────────────────────────────────────────────────
def load_shilla() -> list[dict]:
    rows = []
    with open(SHILLA_CSV, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            name = (r.get("위스키명") or "").strip()
            if not name or is_bundle_noise(name):
                continue
            vol = extract_volume_ml(name)
            usd = _f(r.get("표시가_USD")) or _f(r.get("할인가_USD"))
            if vol is None or usd is None:
                continue
            k1, k2 = _make_key(name)
            rows.append({
                "_k1": k1, "_k2": k2, "_vol": vol, "_usd": usd,
                "name": name, "할인가_USD": _f(r.get("할인가_USD")),
                "정상가_USD": _f(r.get("정상가_USD")), "할인율_%": r.get("할인율_%"),
            })
    return rows


def load_lotte() -> list[dict]:
    rows = []
    with open(LOTTE_CSV, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            name = (r.get("name") or "").strip()
            if not name or is_bundle_noise(name):
                continue
            vol = None
            try:
                vol = int(r.get("volume_ml") or 0) or None
            except ValueError:
                vol = None
            if vol is None:
                vol = extract_volume_ml(name)
            usd = _f(r.get("sale_price"))
            if vol is None or usd is None:
                continue
            k1, k2 = _make_key(name)
            rows.append({
                "_k1": k1, "_k2": k2, "_vol": vol, "_usd": usd,
                "name": name, "할인가_USD": usd,
                "정상가_USD": _f(r.get("regular_price")), "할인율_%": r.get("discount_pct"),
            })
    return rows


def load_ssg() -> list[dict]:
    rows = []
    with open(SSG_CSV, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            name = (r.get("name") or "").strip()
            if not name or is_bundle_noise(name):
                continue
            vol = None
            try:
                vol = int(r.get("volume_ml") or 0) or None
            except ValueError:
                vol = None
            if vol is None:
                vol = extract_volume_ml(name)
            usd = _f(r.get("sale_price"))
            if vol is None or usd is None:
                continue
            k1, k2 = _make_key(name)
            rows.append({
                "_k1": k1, "_k2": k2, "_vol": vol, "_usd": usd,
                "name": name, "할인가_USD": usd,
                "정상가_USD": _f(r.get("regular_price")), "할인율_%": r.get("discount_pct"),
            })
    return rows


def per100(row: dict) -> float:
    return row["_usd"] / row["_vol"] * 100.0


# ── 국내최저가(데일리샷+트레이더스+코스트코) 매칭 ───────────────────────────────
def _domestic_label(market: str, channel: str) -> str | None:
    """normalized 행의 market/channel → 보드 지정 소스 라벨. 비대상이면 None.
    (analyze_attractiveness.load_domestic 의 라벨링과 동일 규칙, 3소스만 통과.)"""
    if market == "KR-DS":
        return "데일리샷"
    ch = channel or ""
    if "코스트코" in ch or "costco" in ch.lower():
        return "코스트코"
    if "트레이더스" in ch:
        return "트레이더스"
    return None  # 이마트/롯데마트/기타 마트는 보드 지정 3소스 밖 → 제외


def load_domestic_floor() -> dict:
    """canonical_id → {floor, source, prev, date, vol} (국내최저가, 보드 3소스 한정).

    floor = 소스별 '최신 수집일 가격' 중 최소값(CMPA-496 per_source_latest_floor 재사용).
    같은 소스의 과거 저가는 superseded — 단순 min() 금지.
    vol = floor 가격이 관측된 실제 용량(ml). canonical _vol(흔히 700 일반값)이 아닌 실관측 용량을
    써야 100ml당 단가가 정확하다(CMPA-666 보드 지적: 1L·750ml 제품의 700 분모 왜곡 방지).
    """
    from collections import defaultdict
    obs = defaultdict(list)            # cid -> [(source, date, price)]
    volrows = defaultdict(list)        # (cid, source) -> [(date, price, vol)]
    latest = {}                        # cid -> 최신 수집일(메타 표기용)
    with open(NORMALIZED_CSV, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            if r.get("status") != "matched":
                continue
            label = _domestic_label(r.get("market", ""), r.get("channel", ""))
            if label not in DOMESTIC_SOURCES:
                continue
            try:
                price = float(r["price_krw"])
            except (ValueError, TypeError):
                continue
            if price <= 0:
                continue
            cid = r["canonical_id"]
            date = (r.get("date") or "").strip()
            try:
                vol = int(r.get("volume_ml") or 0) or None
            except (ValueError, TypeError):
                vol = None
            obs[cid].append((label, date, price))
            volrows[(cid, label)].append((date, price, vol))
            if date > latest.get(cid, ""):
                latest[cid] = date
    out = {}
    for cid, ob in obs.items():
        fl = per_source_latest_floor(ob)
        if not fl:
            continue
        floor, source, prev = fl
        # floor 가격이 관측된 실제 용량 찾기: 해당 소스의 최신일 중 floor 가격 행의 용량.
        rows = volrows.get((cid, source), [])
        vol = None
        if rows:
            ld = max(d for d, _, _ in rows)
            for d, p, v in rows:
                if d == ld and abs(p - floor) < 1 and v:
                    vol = v
                    break
        out[cid] = {"floor": floor, "source": source, "prev": prev,
                    "date": latest.get(cid, ""), "vol": vol}
    return out


def _load_canon_index() -> list:
    """canonical 리스트(_norm·_vol·id·name_ko). 가장 긴(구체적) 부분일치 매칭용."""
    canon = load_canonical()
    # 긴 이름 우선 매칭(더 구체적 제품이 먼저 잡히게)
    return sorted(canon, key=lambda c: -len(c["_norm"]))


# 듀오팩/멀티팩(서로 다른 술 2병 묶음) — 100ml당 단가가 무의미해 국내최저 비교 제외.
_MULTIPACK_PAT = re.compile(r"듀오|듀얼|\bduo\b|x\s*2|×\s*2|2\s*병|세트팩")


def canon_for(name: str, canon_index: list):
    """제품명 → canonical 행(부분일치 + CMPA-177 토큰 가드). 없으면 None."""
    if _MULTIPACK_PAT.search((name or "").lower()):
        return None  # 멀티팩은 단일 canonical 매칭이 부정확 → 국내최저 enrich 생략
    nm = norm(name)
    for c in canon_index:
        cn = c["_norm"]
        if not cn or cn not in nm:
            continue
        if not _cmpa177_ok(nm, cn):   # 년수/CS/피티드/셰리/버번 비대칭이면 다른 제품
            continue
        # 모델/프루프 익스프레션 가드(CMPA-669): 표준 '메이커스 마크'(canonical)가 '메이커스 마크 46/101'
        # 같은 구체적 익스프레션을 부분일치로 삼키는 오매칭 차단. 우리 이름에 canonical(이름+나이)에
        # 없는 결정 숫자가 있으면 다른 제품 → 제외. (나이만 다른 대표표기 차이는 _cmpa177_ok 가 허용.)
        canon_raw = f"{c.get('name_ko') or ''} {c.get('name_en') or ''}"
        allowed = set(_model_nums(canon_raw))
        try:
            if c.get("age"):
                allowed.add(int(c["age"]))
        except (ValueError, TypeError):
            pass
        if _model_nums(name) - allowed:
            continue
        return c
    return None


# ── 다단계 매처 ────────────────────────────────────────────────────────────────
def _build_index(rows: list[dict], key_field: str) -> dict[str, list[dict]]:
    """key_field → [row, ...] 인덱스 구축. 중복 키(같은 제품 여러 행)도 지원."""
    idx: dict[str, list[dict]] = {}
    for r in rows:
        k = r[key_field]
        idx.setdefault(k, []).append(r)
    return idx


def _best_row(candidates: list[dict], target_vol: int | None = None) -> dict | None:
    """복수 후보 중 용량이 가장 가까운 것 반환. target_vol 없으면 첫 번째."""
    if not candidates:
        return None
    if target_vol is None or len(candidates) == 1:
        return candidates[0]
    return min(candidates, key=lambda r: abs(r["_vol"] - target_vol))


def match_row(anchor: dict, pool_k1: dict, pool_k2: dict,
              consumed_k1: set | None = None) -> dict | None:
    """anchor row에 대해 pool에서 최선 매치를 찾는다.
    Stage1: _k1 완전 일치
    Stage2: _k2 완전 일치 + CMPA-177 가드
    consumed_k1: 이미 다른 앵커가 독점한 pool 행의 k1 집합 (중복 매칭 방지)
    """
    ak1 = anchor["_k1"]
    ak2 = anchor["_k2"]
    av = anchor["_vol"]

    def _not_consumed(c: dict) -> bool:
        return consumed_k1 is None or c["_k1"] not in consumed_k1

    # Stage1: 완전 일치
    if ak1 in pool_k1:
        cands = [c for c in pool_k1[ak1]
                 if _vol_compatible(c["_vol"], av) and _not_consumed(c)]
        best = _best_row(cands, av)
        if best:
            return best

    # Stage2: 변환 키 일치 + CMPA-177
    if ak2 in pool_k2:
        cands = []
        for c in pool_k2[ak2]:
            if (_vol_compatible(c["_vol"], av)
                    and _cmpa177_ok(ak2, c["_k2"])
                    and _not_consumed(c)):
                cands.append(c)
        best = _best_row(cands, av)
        if best:
            return best

    return None


def build_rows():
    shilla = load_shilla()
    lotte = load_lotte()
    ssg = load_ssg()

    # 각 shop의 k1/k2 인덱스 구축
    l_k1 = _build_index(lotte, "_k1")
    l_k2 = _build_index(lotte, "_k2")
    g_k1 = _build_index(ssg, "_k1")
    g_k2 = _build_index(ssg, "_k2")

    # 신라를 앵커로 매칭
    # consumed: 이미 다른 신라 행이 독점한 롯데/신세계 pool 행 k1 집합
    consumed_lotte: set[str] = set()
    consumed_ssg: set[str] = set()
    matched_lotte: set[str] = set()    # 이미 매칭된 롯데 행 (중복 방지)
    matched_ssg: set[str] = set()      # 이미 매칭된 신세계 행 (중복 방지)
    matched_shilla: set[str] = set()   # 롯데+신세계 교집합으로 신라에 매칭된 것

    both = []
    shilla_only_rows = []

    for s in shilla:
        l = match_row(s, l_k1, l_k2, consumed_lotte)
        g = match_row(s, g_k1, g_k2, consumed_ssg)

        present = sum(1 for x in (s, l, g) if x is not None)

        if present >= 2:
            sp = per100(s)
            lp = per100(l) if l else None
            gp = per100(g) if g else None

            prices = {k: v for k, v in {"신라": sp, "롯데": lp, "신세계": gp}.items() if v is not None}
            min_price = min(prices.values())
            max_price = max(prices.values())
            cheapest_shops = [k for k, v in prices.items() if v == min_price]
            cheapest = "/".join(cheapest_shops)
            spread_usd = round(max_price - min_price, 2)
            spread_pct = round((max_price - min_price) / max_price * 100.0, 1) if max_price else 0.0

            if sp is not None and lp is not None:
                sl_cheaper = "롯데" if lp < sp else ("신라" if sp < lp else "동일")
                sl_diff = round(abs(lp - sp), 2)
            else:
                sl_cheaper = ""
                sl_diff = None

            vols = [x["_vol"] for x in (s, l, g) if x is not None]
            vol_display = vols[0] if len(set(vols)) == 1 else "/".join(str(v) for v in vols)

            both.append({
                "제품명": s["name"],   # 신라 표기가 가장 완성도 높음
                "용량ml": vol_display,
                "신라용량ml": s["_vol"],
                "롯데용량ml": l["_vol"] if l else "",
                "신세계용량ml": g["_vol"] if g else "",
                "신라_USD": round(s["_usd"], 2),
                "신라_USD100ml": round(sp, 2),
                "롯데_USD": round(l["_usd"], 2) if l else "",
                "롯데_USD100ml": round(lp, 2) if lp is not None else "",
                "신세계_USD": round(g["_usd"], 2) if g else "",
                "신세계_USD100ml": round(gp, 2) if gp is not None else "",
                "최저가_면세점": cheapest,
                "가격차_USD100ml": spread_usd,
                "가격차_%": spread_pct,
                "더싼곳": sl_cheaper,
                "차이_USD100ml": sl_diff if sl_diff is not None else "",
                "신라상품명": s["name"],
                "롯데상품명": l["name"] if l else "",
                "신세계상품명": g["name"] if g else "",
            })

            if l:
                matched_lotte.add(l["_k1"])
                consumed_lotte.add(l["_k1"])
            if g:
                matched_ssg.add(g["_k1"])
                consumed_ssg.add(g["_k1"])
            matched_shilla.add(s["_k1"])
        else:
            shilla_only_rows.append(s)

    # 롯데·신세계에만 있는 제품: 신라 앵커 이후 남은 것들
    # 롯데 단독 (신세계와도 매칭 안된 것)
    lotte_unmatched = [r for r in lotte if r["_k1"] not in matched_lotte]
    ssg_unmatched = [r for r in ssg if r["_k1"] not in matched_ssg]

    # 롯데 ↔ 신세계 추가 교집합 (신라 없는 것)
    lg_k1 = _build_index(lotte_unmatched, "_k1")
    lg_k2 = _build_index(lotte_unmatched, "_k2")
    lg_matched_ssg: set[str] = set()
    lg_matched_lotte: set[str] = set()
    lg_consumed_lotte: set[str] = set()

    for g in ssg_unmatched:
        l = match_row(g, lg_k1, lg_k2, lg_consumed_lotte)
        if l:
            gp = per100(g)
            lp = per100(l)
            prices = {"롯데": lp, "신세계": gp}
            min_price = min(prices.values())
            max_price = max(prices.values())
            cheapest_shops = [k for k, v in prices.items() if v == min_price]
            spread_usd = round(max_price - min_price, 2)
            spread_pct = round((max_price - min_price) / max_price * 100.0, 1) if max_price else 0.0

            vols = [g["_vol"], l["_vol"]]
            vol_display = vols[0] if len(set(vols)) == 1 else "/".join(str(v) for v in vols)

            both.append({
                "제품명": g["name"],
                "용량ml": vol_display,
                "신라용량ml": "",
                "롯데용량ml": l["_vol"],
                "신세계용량ml": g["_vol"],
                "신라_USD": "",
                "신라_USD100ml": "",
                "롯데_USD": round(l["_usd"], 2),
                "롯데_USD100ml": round(lp, 2),
                "신세계_USD": round(g["_usd"], 2),
                "신세계_USD100ml": round(gp, 2),
                "최저가_면세점": "/".join(cheapest_shops),
                "가격차_USD100ml": spread_usd,
                "가격차_%": spread_pct,
                "더싼곳": "",
                "차이_USD100ml": "",
                "신라상품명": "",
                "롯데상품명": l["name"],
                "신세계상품명": g["name"],
            })
            lg_matched_ssg.add(g["_k1"])
            lg_matched_lotte.add(l["_k1"])
            lg_consumed_lotte.add(l["_k1"])

    # 신라단독 / 롯데단독 / 신세계단독
    shilla_only = []
    for s in shilla_only_rows:
        if s["_k1"] not in matched_shilla:
            shilla_only.append({
                "제품명": s["name"], "용량ml": s["_vol"],
                "신라_USD": round(s["_usd"], 2), "신라_USD100ml": round(per100(s), 2),
                "신라상품명": s["name"],
            })

    lotte_only = []
    for l in lotte_unmatched:
        if l["_k1"] not in lg_matched_lotte:
            lotte_only.append({
                "제품명": l["name"], "용량ml": l["_vol"],
                "롯데_USD": round(l["_usd"], 2), "롯데_USD100ml": round(per100(l), 2),
                "롯데상품명": l["name"],
            })

    ssg_only = []
    for g in ssg_unmatched:
        if g["_k1"] not in lg_matched_ssg:
            ssg_only.append({
                "제품명": g["name"], "용량ml": g["_vol"],
                "신세계_USD": round(g["_usd"], 2), "신세계_USD100ml": round(per100(g), 2),
                "신세계상품명": g["name"],
            })

    # 국내최저가(데일리샷+트레이더스+코스트코) enrich
    fx, fx_asof = load_fx()
    canon_index = _load_canon_index()
    dom_floor = load_domestic_floor()
    online = load_online_cache()
    dom_meta = {"fx": fx, "fx_asof": fx_asof,
                "dates": sorted({v["date"] for v in dom_floor.values() if v["date"]}),
                "online_priced": sum(1 for v in online.values() if v.get("price"))}
    for grp in (both, shilla_only, lotte_only, ssg_only):
        for r in grp:
            _enrich_domestic(r, canon_index, dom_floor, fx, online)

    # 가격차 큰 순 정렬
    both.sort(key=lambda x: -(x["가격차_USD100ml"] or 0))
    return (both, shilla_only, lotte_only, ssg_only,
            len(shilla), len(lotte), len(ssg), dom_meta)


# 면세 100ml당 USD 컬럼들(존재 시 최저가 산출에 사용)
_DUTY_P100_FIELDS = ("신라_USD100ml", "롯데_USD100ml", "신세계_USD100ml")

# 데일리샷 온라인 보강 캐시(enrich_online_dailyshot.py 산출) — 워치리스트 밖 제품 커버.
# 하드코딩 날짜 대신 최신 _dailyshot_compare_<date>.csv 자동탐색(CMPA-672). 없어도 비치명.
ONLINE_CACHE, ONLINE_DATE = _latest_snapshot(
    os.path.join(ROOT, "data", "whisky-prices", "_dailyshot_compare_????-??-??.csv"),
    required=False)


# 미니어처/샘플 오매칭 가드: 700ml 질의가 '…미니어처/미니/샘플'에 붙으면 가짜 저가가 된다
# (예: '글렌알라키 8년 700ml'→'글렌알라키 8년 미니어처' ₩25,500). ds_name 으로 차단.
_MINI_PAT = re.compile(r"미니어처|미니츄어|미니어쳐|샘플|\bmini\b|miniature|sample|\d{2,3}\s*ml\s*세트")


def load_online_cache() -> dict:
    """norm(제품명) → {price, vol, seller} (데일리샷 온라인 전국 최저, 면세·미니어처 제외)."""
    out = {}
    if not os.path.exists(ONLINE_CACHE):
        return out
    with open(ONLINE_CACHE, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            try:
                price = float(r.get("ds_price_krw") or 0)
            except (ValueError, TypeError):
                price = 0
            if price <= 0:
                continue
            # 미니어처/샘플로 매칭된 행은 본품 가격이 아니므로 제외(오매칭 가짜 저가 방지).
            ds_name = r.get("ds_name") or ""
            if _MINI_PAT.search(ds_name.lower()):
                continue
            # 익스프레션 오매칭 가드(CMPA-669): 데일리샷이 같은 브랜드 다른 익스프레션에 붙은 경우 제외.
            # 예: '와일드터키 켄터키 스피릿'→'와일드 터키 101', '메이커스 마크 101'→'메이커스 마크 46'.
            our_name = r.get("제품명") or ""
            if not _proof_ok(our_name, ds_name):
                continue
            try:
                vol = int(r.get("ds_vol_ml") or 0) or None
            except (ValueError, TypeError):
                vol = None
            out[r["_k1"]] = {"price": price, "vol": vol, "seller": r.get("ds_seller", ""),
                             "item_id": (r.get("ds_item_id") or "").strip()}
    return out


STORY_CSV = os.path.join(ROOT, "data", "shilla-dutyfree", "whisky-story.csv")


def load_story_lookup() -> dict:
    """제품명 → 스토리/맛 레코드. norm 키 + 용량 제거 키 두 단계로 색인(있으면 보여주기용)."""
    exact, novol = {}, {}
    if not os.path.exists(STORY_CSV):
        return {"exact": exact, "novol": novol}
    vol_pat = re.compile(r"\d+(?:ml|l)$")
    with open(STORY_CSV, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            nm = (r.get("위스키명") or "").strip()
            if not nm:
                continue
            k = norm(nm)
            exact.setdefault(k, r)
            nv = vol_pat.sub("", k)
            # 용량 제거 키는 충돌 시 첫 항목만(모호하면 안 쓰도록 표시)
            if nv in novol and novol[nv] is not None and norm((novol[nv].get("위스키명") or "")) != k:
                novol[nv] = None  # 모호 → 비활성
            else:
                novol.setdefault(nv, r)
    return {"exact": exact, "novol": novol}


def story_for(name: str, story_idx: dict) -> dict | None:
    """제품명에 맞는 스토리 레코드. exact(norm) 우선, 없으면 용량 제거 키(모호하지 않을 때)."""
    k = norm(name)
    rec = story_idx["exact"].get(k)
    if rec:
        return rec
    vol_pat = re.compile(r"\d+(?:ml|l)$")
    nv = vol_pat.sub("", k)
    rec = story_idx["novol"].get(nv)
    if rec and _cmpa177_ok(k, norm(rec.get("위스키명") or "")):
        return rec
    return None


def _cell_safe(t: str) -> str:
    """표 셀 안전(파이프·줄바꿈 제거)."""
    return (t or "").replace("|", "/").replace("\n", " ").replace("\r", " ").strip()


def story_cell_html(name: str, story_idx: dict) -> str:
    """제품 셀. 스토리/맛이 있으면 제품명 클릭 시 펼침(<details>), 없으면 제품명 텍스트.
    (raw HTML <table> 셀 안에서 렌더 — kramdown 파이프표는 블록HTML 이스케이프되므로 표를 HTML로 짠다.)"""
    rec = story_for(name, story_idx)
    if not rec:
        return name
    head = " · ".join(_cell_safe(rec.get(k)) for k in ("증류소", "지역", "도수", "캐스크")
                      if (rec.get(k) or "").strip())
    parts = []
    if head:
        parts.append(f"<b>{head}</b>")
    taste = _cell_safe(rec.get("맛_노트"))
    if taste:
        parts.append(f"🥃 {taste}")
    story = _cell_safe(rec.get("스토리"))
    if story:
        parts.append(f"📖 {story}")
    src = _cell_safe(rec.get("출처"))
    if src:
        parts.append(f"<small>출처: {src}</small>")
    body = "<br>".join(parts)
    return (f'<details><summary style="cursor:pointer">{name}</summary>'
            f'<div style="font-weight:400;font-size:.92em;line-height:1.55;'
            f'padding:6px 2px 2px;white-space:normal;max-width:min(78vw,440px)">'
            f"{body}</div></details>")


def dom_src_html(src: str, name: str, tpid: str) -> str:
    """국내 소매가 셀의 출처 표기. (데일리샷)은 제품 페이지(또는 검색) 링크·새창,
    트레이더스/코스트코는 🛒 텍스트(물리 매장이라 상품 링크 없음)."""
    import urllib.parse as ulib
    tpid = (tpid or "").strip()
    if src == "데일리샷" and tpid:
        url = f"https://dailyshot.co/m/item/{tpid}"
        return f'<a href="{url}" target="_blank" rel="noopener">({src})</a>'
    if src == "데일리샷":
        url = f"https://dailyshot.co/m/search/result?q={ulib.quote(name)}"
        return f'<a href="{url}" target="_blank" rel="noopener">({src})</a>'
    if src in ("트레이더스", "코스트코"):
        return f"🛒 ({src})"
    return f"({src})"


def _enrich_domestic(row: dict, canon_index: list, dom_floor: dict, fx: float,
                     online: dict | None = None):
    """비교/단독 행에 국내최저가(₩) + 면세 vs 국내 차이를 채운다.

    국내최저가 후보 = ① 워치리스트 floor(데일리샷+트레이더스+코스트코, canonical 표준용량)
    ② 데일리샷 온라인 보강(enrich_online_dailyshot, 실제 ds 용량) — 둘 중 100ml당 최소.
    면세최저_₩100ml = 행에 있는 면세 100ml당 USD 최저 × 단일 환율(fx).
    면세vs국내_% = (국내 - 면세최저)/국내 × 100  (양수=면세가 국내보다 그만큼 싸다)
    """
    for c in ("국내최저_₩", "국내최저_₩100ml", "국내최저_소스", "국내수집일", "국내최저_용량ml",
              "국내최저_tpid", "면세최저_₩100ml", "면세vs국내_%", "최저처(면세vs국내)"):
        row.setdefault(c, "")
    # 미니(<350ml)는 100ml당 단가가 왜곡돼 국내최저 비교 부적합(canonical MINI_ML 일관)
    vols = []
    for vf in ("신라용량ml", "롯데용량ml", "신세계용량ml", "용량ml"):
        v = row.get(vf)
        try:
            iv = int(v)
            if iv > 0:
                vols.append(iv)
        except (ValueError, TypeError):
            continue
    if vols and min(vols) < 350:
        return
    cands = []  # (p100, krw_bottle, source, date, tpid, real_vol)
    # real_vol = 실관측 용량(없으면 None — 표시에 가짜 용량 안 씀). 100ml당 계산은 vol_calc(추정 포함).
    c = canon_for(row.get("제품명", ""), canon_index)
    if c:
        fl = dom_floor.get(c["id"])
        if fl:
            real_vol = fl.get("vol")          # 실관측 용량(없으면 None)
            vol_calc = real_vol or c.get("_vol") or 700
            cands.append((fl["floor"] / vol_calc * 100.0, fl["floor"], fl["source"],
                          fl["date"], "", real_vol))
    if online:
        oc = online.get(norm(row.get("제품명", "")))
        if oc:
            real_vol = oc["vol"]              # 데일리샷 ds_vol_ml(없으면 None)
            vol_calc = real_vol or (c.get("_vol") if c else None) or 700
            cands.append((oc["price"] / vol_calc * 100.0, oc["price"], "데일리샷", ONLINE_DATE,
                          oc.get("item_id", ""), real_vol))
    if not cands:
        return
    # 오매칭 백스톱(CLAUDE.md ~2.5배, 양방향): 국내 100ml당가가 면세 최저의
    #  ① 2.5배 이상(비싼 제품에 오매칭 — 예: '글렌파클라스 105'→'글렌파클라스 25년') 또는
    #  ② 1/2.5 이하(싼/작은 제품에 오매칭 — 예: '와일드터키 마스터스 킵'→'와일드터키 101' ₩39,900,
    #     '글렌모렌지 오리지널'→미니/오류 ₩12,000)이면 같은 제품이 아닐 가능성이 높다 → 제외.
    # 면세는 무관세라 같은 SKU 국내가가 면세의 40% 미만일 수는 없다(있으면 오매칭).
    duty_p100 = [row[f] for f in _DUTY_P100_FIELDS if row.get(f) not in ("", None)]
    duty_min_krw = (min(duty_p100) * fx) if duty_p100 else None
    if duty_min_krw:
        cands = [cc for cc in cands
                 if duty_min_krw / 2.5 <= cc[0] < 2.5 * duty_min_krw]
    if not cands:
        return
    dom_p100, dom_krw, dom_src, dom_date, dom_tpid, dom_vol = min(cands, key=lambda x: x[0])
    row["국내최저_₩"] = round(dom_krw)
    row["국내최저_₩100ml"] = round(dom_p100)
    row["국내최저_소스"] = dom_src
    row["국내수집일"] = dom_date
    row["국내최저_tpid"] = dom_tpid
    row["국내최저_용량ml"] = dom_vol or ""

    if not duty_p100:
        return
    row["면세최저_₩100ml"] = round(duty_min_krw)
    if dom_p100 > 0:
        gap = (dom_p100 - duty_min_krw) / dom_p100 * 100.0
        row["면세vs국내_%"] = round(gap, 1)
        row["최저처(면세vs국내)"] = "면세" if duty_min_krw < dom_p100 else "국내"


# ── 엑셀 작성 ────────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
META_FILL = PatternFill("solid", fgColor="F2F2F2")


def _write_sheet(ws, rows, cols):
    ws.append(cols)
    for cell in ws[1]:
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
    for i, c in enumerate(cols, 1):
        w = max(len(str(c)) * 2, *(len(str(r.get(c, ""))) for r in rows)) if rows else len(c) * 2
        ws.column_dimensions[get_column_letter(i)].width = min(max(w + 2, 10), 40)
    ws.freeze_panes = "A2"


def write_xlsx(path, both, shilla_only, lotte_only, ssg_only, n_shilla, n_lotte, n_ssg,
               dom_meta):
    wb = openpyxl.Workbook()
    meta = wb.active
    meta.title = "메타"

    all3 = sum(1 for r in both if r["신라_USD"] != "" and r["롯데_USD"] != "" and r["신세계_USD"] != "")
    sl_only_pair = sum(1 for r in both if r["신세계_USD"] == "" and r["신라_USD"] != "" and r["롯데_USD"] != "")
    sg_only_pair = sum(1 for r in both if r["롯데_USD"] == "" and r["신라_USD"] != "" and r["신세계_USD"] != "")
    lg_only_pair = sum(1 for r in both if r["신라_USD"] == "" and r["롯데_USD"] != "" and r["신세계_USD"] != "")

    meta_rows = [
        ["면세점 100ml당 가격 비교 (신라 vs 롯데 vs 신세계)", ""],
        ["기준 통화", "USD (면세 USD 표기 — FX 노이즈 없음)"],
        ["롯데 수집일", LOTTE_DATE],
        ["신라 수집일", SHILLA_DATE],
        ["신세계(SSG) 수집일", SSG_DATE],
        ["가격 기준", "롯데/신세계=sale_price(실판매가) · 신라=표시가_USD(파이프라인 1차가)"],
        ["100ml당 가격", "가격 / volume_ml × 100"],
        ["국내최저가", "데일리샷 + 트레이더스 + 코스트코 합산 최저(보드 지정 3소스). 소스별 최신가 중 최소."],
        ["국내최저가 수집일", " · ".join(dom_meta["dates"]) or "(데이터 없음)"],
        ["환율(면세₩ 환산)", f"1 USD = {dom_meta['fx']:,.2f} KRW (asof {dom_meta['fx_asof']}, 단일 환율)"],
        ["면세vs국내_%", "양수 = 면세 100ml당가가 국내최저보다 그만큼 더 싸다(면세 메리트)."],
        ["매칭 방법", "다단계 직접 이름 매처 (Stage1: 정규화 완전일치, Stage2: 변형 통일 후 재매칭, CMPA-177 가드)"],
        ["비교(교집합) 종수", len(both)],
        ["  → 3개 면세점 모두", all3],
        ["  → 신라+롯데만", sl_only_pair],
        ["  → 신라+신세계만", sg_only_pair],
        ["  → 롯데+신세계만", lg_only_pair],
        ["신라단독 / 롯데단독 / 신세계단독", f"{len(shilla_only)} / {len(lotte_only)} / {len(ssg_only)}"],
        ["신라/롯데/신세계 매칭후보 행수", f"{n_shilla} / {n_lotte} / {n_ssg}"],
        ["주의", "면세가는 출국·면세한도(2병/2L/$400) 조건. '수집일 기준'값(현재가 단정 아님)."],
    ]
    for row in meta_rows:
        meta.append(row)
    meta["A1"].font = Font(bold=True, size=13)
    for r in range(2, len(meta_rows) + 1):
        meta.cell(r, 1).font = Font(bold=True)
        meta.cell(r, 1).fill = META_FILL
    meta.column_dimensions["A"].width = 28
    meta.column_dimensions["B"].width = 70

    cmp_cols = [
        "제품명", "용량ml",
        "신라용량ml", "롯데용량ml", "신세계용량ml",
        "신라_USD", "신라_USD100ml",
        "롯데_USD", "롯데_USD100ml",
        "신세계_USD", "신세계_USD100ml",
        "최저가_면세점", "가격차_USD100ml", "가격차_%",
        # 국내최저가(데일리샷+트레이더스+코스트코) 비교
        "국내최저_₩", "국내최저_₩100ml", "국내최저_소스", "국내수집일",
        "면세최저_₩100ml", "면세vs국내_%", "최저처(면세vs국내)",
        "더싼곳", "차이_USD100ml",
        "신라상품명", "롯데상품명", "신세계상품명",
    ]
    dom_cols = ["국내최저_₩", "국내최저_₩100ml", "국내최저_소스", "국내수집일",
                "면세최저_₩100ml", "면세vs국내_%", "최저처(면세vs국내)"]
    _write_sheet(wb.create_sheet("비교"), both, cmp_cols)
    _write_sheet(wb.create_sheet("신라단독"), shilla_only,
                 ["제품명", "용량ml", "신라_USD", "신라_USD100ml"] + dom_cols + ["신라상품명"])
    _write_sheet(wb.create_sheet("롯데단독"), lotte_only,
                 ["제품명", "용량ml", "롯데_USD", "롯데_USD100ml"] + dom_cols + ["롯데상품명"])
    _write_sheet(wb.create_sheet("신세계단독"), ssg_only,
                 ["제품명", "용량ml", "신세계_USD", "신세계_USD100ml"] + dom_cols + ["신세계상품명"])
    wb.save(path)
    return cmp_cols


def write_csv(path, both, cols):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        w.writerows(both)


# ── 메인 텍스트 캐러셀(면세 vs 국내 'gap' 하이라이트) — CMPA-693 보드 2026-06-29 ──
#   지표 gap = 국내최저_₩100ml − 면세최저_₩100ml (+면세 이득 / −소매 이득).
#   인기 밴드(병당[면세] 50,000~300,000) 안에서 TOP(면세 이득)·BOTTOM(소매 이득)·
#   전일대비 큰 변동(+원인 귀속)·스코어보드(면세 최저가 챔피언)를 한 줄씩 뽑아
#   compare 글 front matter 의 `carousel:` 리스트로 적재 → 홈 _HOME_CAROUSEL 이 회전 노출.
#   입력 = 최신 reports/dutyfree-compare/면세점_100ml_가격비교_<date>.csv + 전일분(movers).
CAROUSEL_BAND_LO = 50000     # 병당(면세) 인기 밴드 하한(원)
CAROUSEL_BAND_HI = 300000    # 병당(면세) 인기 밴드 상한(원)
COMPARE_DIR = os.path.join(ROOT, "reports", "dutyfree-compare")
_COMPARE_DATE_RE = re.compile(r"면세점_100ml_가격비교_(\d{4}-\d{2}-\d{2})\.csv$")


def _compare_csv_path(date_str):
    return os.path.join(COMPARE_DIR, f"면세점_100ml_가격비교_{date_str}.csv")


def _read_compare_csv(path):
    try:
        with open(path, encoding="utf-8-sig") as f:
            return list(csv.DictReader(f))
    except (FileNotFoundError, OSError):
        return []


def _prev_compare_date(date_str):
    """date_str 직전(가장 가까운 과거) 비교 CSV 날짜. 없으면 None."""
    import glob
    cand = []
    for p in glob.glob(os.path.join(COMPARE_DIR, "면세점_100ml_가격비교_*.csv")):
        m = _COMPARE_DATE_RE.search(p)
        if m and m.group(1) < date_str:
            cand.append(m.group(1))
    return max(cand) if cand else None


def _cnum(v):
    try:
        s = str(v).replace(",", "").strip()
        return float(s) if s not in ("", "None") else None
    except (TypeError, ValueError):
        return None


def _gap_row(r):
    """비교 CSV 한 행 → gap 계산 dict(양쪽가격·용량 있을 때만). 없으면 None."""
    dom = _cnum(r.get("국내최저_₩100ml"))
    duty = _cnum(r.get("면세최저_₩100ml"))
    vol = _cnum(r.get("용량ml"))
    if dom is None or duty is None or not vol:
        return None
    return {
        "name": r.get("제품명", ""),
        "gap": dom - duty,                       # +면세 이득 / −소매 이득 (100ml당, 정렬키)
        "dom": dom, "duty": duty, "vol": vol,
        "bottle_duty": duty * vol / 100.0,        # 병당(면세) 환산가
        "dom_bottle": _cnum(r.get("국내최저_₩")),   # 병당(국내 소매 실판매가, 앵커)
        "src": r.get("국내최저_소스", "") or "국내",
        "pct": _cnum(r.get("면세vs국내_%")),       # gap/국내×100 = 면세 절약률
        "best_shop": (r.get("최저가_면세점", "") or "").strip(),
    }


# 캐러셀 표시용 제품명 정리 — 한 줄 가독성(보드 2026-06-29 '문구' 피드백).
#   꼬리 용량(700ml)·일반 분류어(블렌디드 스카치 위스키 등)·머리 'NEW' 를 떼어
#   '로얄살루트 21년 빈티지 블렌디드 스카치 위스키 700ml' → '로얄살루트 21년 빈티지'.
#   ⚠️ 숙성년수·CS·피티드·캐스크·에디션(CMPA-177)은 보존 — 제품 구분 토큰이라 삭제 금지.
_VOL_TAIL = re.compile(r"\s*\d+(?:\.\d+)?\s*(?:ml|l)\s*$", re.I)
_GENERIC_TAIL = re.compile(r"\s*(블렌디드\s*)?(싱글\s*몰트\s*)?스카치\s*위스키\s*$")


def _disp_name(name):
    s = re.sub(r"^\s*NEW\s+", "", name)
    s = _VOL_TAIL.sub("", s)
    s = _GENERIC_TAIL.sub("", s)
    s = re.sub(r"\s+위스키\s*$", "", s)
    return s.strip() or name.strip()


def _man(v):
    """원 금액 → '16만원'·'13.9만원'(천원 단위 반올림, 사람이 읽는 병당 표기)."""
    m = round(v / 10000.0, 1)
    return f"{int(m)}만원" if m == int(m) else f"{m:g}만원"


def _md(date_str):
    """YYYY-MM-DD → 'M/D'(작은 날짜 칩, 기사 느낌)."""
    try:
        return f"{int(date_str[5:7])}/{int(date_str[8:10])}"
    except (ValueError, IndexError):
        return date_str


def _traders_ocr_signals():
    """최신 youtube_ocr CSV에서 트레이더스 신규 입고·가격 하락 시그널.

    반환: (total, new_items, drops, latest_date)
      total       — 최신 수집일 총 종수
      new_items   — [(name, price, store)]  이전 수집일 없던 항목
      drops       — [(name, old_price, new_price, pct)]  가격 하락(≥3%) 항목
      latest_date — 최신 수집일 문자열(YYYY-MM-DD) 또는 ""
    """
    try:
        import glob as _glob
        files = sorted(_glob.glob(
            os.path.join(ROOT, "data", "whisky-prices", "????-??_youtube_ocr.csv")))
        if not files:
            return 0, [], [], ""
        rows = []
        for path in files[-2:]:   # 최신 2개월 파일 읽기(월 경계 대비)
            with open(path, encoding="utf-8-sig", newline="") as f:
                for r in csv.DictReader(f):
                    rows.append(r)
        if not rows:
            return 0, [], [], ""
        dates = sorted(set(r.get("가져온날짜", "") for r in rows if r.get("가져온날짜")))
        if not dates:
            return 0, [], [], ""
        latest = dates[-1]
        latest_rows = [r for r in rows if r.get("가져온날짜") == latest]
        prev_rows = {}
        if len(dates) >= 2:
            prev = dates[-2]
            for r in rows:
                if r.get("가져온날짜") == prev:
                    key = r.get("술이름", "").strip()
                    prev_rows[key] = r
        new_items, drops = [], []
        for r in latest_rows:
            name = r.get("술이름", "").strip()
            try:
                price = int(float(r.get("가격_KRW", 0) or 0))
            except (ValueError, TypeError):
                price = 0
            store = r.get("위치", "") or ""
            if name not in prev_rows:
                new_items.append((name, price, store))
            else:
                try:
                    old_price = int(float(prev_rows[name].get("가격_KRW", 0) or 0))
                except (ValueError, TypeError):
                    old_price = 0
                if old_price > 0 and 0 < price < old_price:
                    pct = (old_price - price) / old_price * 100
                    if pct >= 3:
                        drops.append((name, old_price, price, pct))
        drops.sort(key=lambda x: -x[3])
        return len(latest_rows), new_items[:3], drops[:3], latest
    except Exception:
        return 0, [], [], ""


def _retail_signals():
    """데일리샷 최신 스냅샷 기준 소매 가격 하락·신규 HIT 시그널.

    반환: (drops, new_hit)
      drops   — [(name, prev_price, cur_price, delta, pct, acc, loc, url, cf)]
      new_hit — [(name, None, cur_price, None, None, acc, loc, url, cf)]
    데이터 없거나 오류 시 ([], []).
    """
    try:
        _ddc_dir = os.path.join(ROOT, "pipelines", "dailyshot")
        if _ddc_dir not in sys.path:
            sys.path.insert(0, _ddc_dir)
        import detect_dailyshot_changes as _ddc  # noqa: PLC0415
        snaps = _ddc.discover_snapshots()
        info = _ddc.pick_latest_prev(snaps)
        if not info:
            return [], []
        _, prev_path, _, latest_path = info
        prev_rows = _ddc.load_csv(prev_path)
        latest_rows = _ddc.load_csv(latest_path)
        drops, _, new_hit, _ = _ddc.classify(prev_rows, latest_rows)
        return drops, new_hit
    except Exception:
        return [], []


def build_carousel_items(date_str):
    """compare 글 front matter 에 넣을 캐러셀 한 줄 문자열 리스트(설계 §3).

    문구 톤(보드 2026-06-29 피드백): 담백·간결. 'X 싸요'/₩기호/장황한 괄호 대신
    'X 싸다'·'원'·'병값 약 N만원'. 제품명은 _disp_name 으로 용량/분류어 제거.
    """
    cur_rows = _read_compare_csv(_compare_csv_path(date_str))
    if not cur_rows:
        return []
    cur = [g for g in (_gap_row(r) for r in cur_rows) if g]
    band = [g for g in cur if CAROUSEL_BAND_LO <= g["bottle_duty"] <= CAROUSEL_BAND_HI]
    items = []

    md = _md(date_str)

    # 아이템별 랜딩 URL — 면세 관련 = compare 글, 소매 관련 = wprice 소매가 월간 리포트.
    _compare_url = f"/{date_str.replace('-', '/')}/dutyfree-whisky-compare/"
    import glob as _glob
    _wprice_month = date_str[:7]  # "2026-06"
    _wprice_posts = sorted(_glob.glob(
        os.path.join(ROOT, "blog-md", "_posts", f"*-wprice-{_wprice_month}.md")))
    _retail_url = (
        f"/{os.path.basename(_wprice_posts[-1])[:10].replace('-', '/')}/wprice-{_wprice_month}/"
        if _wprice_posts else _compare_url
    )

    # 🟢 면세 이득 TOP 5 — 정렬키=100ml당 gap, 표시=병당+국내 소매가 앵커(보드 확정 2026-06-29).
    #   "{이름} — 국내 소매가 {국내병당}인데 면세 {면세병당}, 약 {절약병당} 싸요 ({pct}% 할인) · {M/D}"
    for g in sorted([x for x in band if x["gap"] > 0], key=lambda x: -x["gap"])[:5]:
        dn = _disp_name(g["name"])
        if g["dom_bottle"] and g["dom_bottle"] > g["bottle_duty"]:
            save = g["dom_bottle"] - g["bottle_duty"]
            pct = f" ({g['pct']:.0f}% 할인)" if g["pct"] else ""
            items.append((
                f"🟢 {dn} — 국내 소매가 {_man(g['dom_bottle'])}인데 면세 "
                f"{_man(g['bottle_duty'])}, 약 {_man(save)} 싸요{pct} · {md}",
                _compare_url))
        else:  # 국내 소매가 앵커 없으면 면세 병당만 담백하게
            pct = f" ({g['pct']:.0f}% 싸요)" if g["pct"] else ""
            items.append((
                f"🟢 {dn} — 면세에서 사면 한 병 {_man(g['bottle_duty'])}, "
                f"국내보다 이득{pct} · {md}",
                _compare_url))

    # 🔴 소매가 더 싸다 BOTTOM 3 — gap<0(국내가 더 쌈), 소매 이득 큰 순.
    #   "{이름} — 면세 {면세병당}인데 국내가 {국내병당}, 약 {차액} 더 싸요 · {M/D}"
    for g in sorted([x for x in band if x["gap"] < 0], key=lambda x: x["gap"])[:3]:
        dn = _disp_name(g["name"])
        if g["dom_bottle"] and g["dom_bottle"] < g["bottle_duty"]:
            diff = g["bottle_duty"] - g["dom_bottle"]
            items.append((
                f"🔴 {dn} — 면세 {_man(g['bottle_duty'])}인데 국내가 "
                f"{_man(g['dom_bottle'])}, 약 {_man(diff)} 더 싸요 · {md}",
                _compare_url))
        else:
            items.append((
                f"🔴 {dn} — 지금은 국내에서 사는 게 면세보다 싸요 · {md}",
                _compare_url))

    # 🔀 오늘 큰 변동 3 — 전일대비 gap 순위 이동 + 원인 귀속
    prev_date = _prev_compare_date(date_str)
    if prev_date:
        prev = {g["name"]: g for g in
                (_gap_row(r) for r in _read_compare_csv(_compare_csv_path(prev_date))) if g}
        cur_rank = {g["name"]: i + 1 for i, g in
                    enumerate(sorted(band, key=lambda x: -x["gap"]))}
        prev_band = sorted([g for g in prev.values()
                            if CAROUSEL_BAND_LO <= g["bottle_duty"] <= CAROUSEL_BAND_HI],
                           key=lambda x: -x["gap"])
        prev_rank = {g["name"]: i + 1 for i, g in enumerate(prev_band)}
        movers = []
        for g in band:
            n = g["name"]
            if n not in prev or n not in prev_rank:
                continue
            d_rank = prev_rank[n] - cur_rank[n]          # +면 순위 상승
            gap_d = g["gap"] - prev[n]["gap"]
            duty_d = g["duty"] - prev[n]["duty"]
            dom_d = g["dom"] - prev[n]["dom"]
            if abs(d_rank) < 1 and abs(gap_d) < 500:
                continue
            movers.append((abs(d_rank), abs(gap_d), d_rank, duty_d, dom_d, g))
        movers.sort(key=lambda x: (-x[0], -x[1]))
        for _, __, d_rank, duty_d, dom_d, g in movers[:3]:
            dn = _disp_name(g["name"])
            # 사람 행동으로 번역(전문어 금지·보드 확정 2026-06-29): 면세에서 사기 좋아졌나/아쉬워졌나.
            if abs(duty_d) >= abs(dom_d):
                msg = ("면세 할인이 줄어 면세에서 사기 살짝 아쉬워졌어요" if duty_d > 0
                       else "면세가 더 싸져서 지금 노릴 만해요")
            else:
                msg = ("국내값이 올라 면세가 상대적으로 더 이득이 됐어요" if dom_d > 0
                       else "국내값이 내려 면세 매력이 조금 줄었어요")
            items.append((f"🔀 {dn}, {msg} · {md}", _compare_url))

        # 🏆 스코어보드 — 쉬운 말(보드 확정): "가장 많은 곳은 OO예요 (어제까진 △△)"
        from collections import Counter
        cnt_cur = Counter(g["best_shop"] for g in cur if g["best_shop"])
        cnt_prev = Counter(g["best_shop"] for g in prev.values() if g["best_shop"])
        if any(cnt_cur.values()):
            top = max(("신세계", "신라", "롯데"), key=lambda s: cnt_cur.get(s, 0))
            n = cnt_cur.get(top, 0)
            chg = ""
            if cnt_prev:
                prev_top = max(("신세계", "신라", "롯데"), key=lambda s: cnt_prev.get(s, 0))
                if any(cnt_prev.values()) and prev_top != top:
                    chg = f" (어제까진 {prev_top}였죠)"
            items.append((
                f"🏆 오늘 면세 최저가가 가장 많은 곳은 {top}예요 ({n}종){chg} · {md}",
                _compare_url))

        # ✈️ 면세 가격 하락 TOP 2 — compare CSV 병당 면세가 하락폭 기준.
        #   >200원/100ml 하락부터 신호.
        duty_drops = []
        for g in band:
            drop_per100 = prev.get(g["name"], {}).get("duty")
            if drop_per100 is None:
                continue
            drop_per100 = prev[g["name"]]["duty"] - g["duty"]  # 양수 = 하락
            if drop_per100 > 200:
                duty_drops.append((drop_per100, g, prev[g["name"]]))
        duty_drops.sort(key=lambda x: -x[0])
        for _, g, pg in duty_drops[:2]:
            dn = _disp_name(g["name"])
            items.append((
                f"✈️ {dn} — 면세가 내렸어요, "
                f"병당 {_man(pg['bottle_duty'])}→{_man(g['bottle_duty'])} · {md}",
                _compare_url))

        # ✈️ 면세 이득 순위권 새 진입 TOP 2 — 어제 밴드에 없던 항목(gap > 0).
        new_entries = sorted(
            [g for g in band if g["gap"] > 0 and g["name"] not in prev_rank],
            key=lambda x: -x["gap"])
        for g in new_entries[:2]:
            dn = _disp_name(g["name"])
            bottle_save = g["gap"] * g["vol"] / 100.0
            items.append((
                f"✈️ {dn} — 오늘 면세 이득 순위권 새로 진입, "
                f"국내보다 약 {_man(bottle_save)} 싸요 · {md}",
                _compare_url))

    # 🆕 소매 신규 입고 TOP 3 — 데일리샷 최신 스냅샷 신규 HIT.
    # 📉 소매 의미 있는 하락 TOP 3 — 데일리샷 최신 스냅샷 가격 하락.
    r_drops, r_new = _retail_signals()
    seen_new = 0
    for nm, _, cur_price, _, _, _, loc, _, cf in r_new:
        if cf or cur_price is None:
            continue
        dn = _disp_name(nm)
        loc_str = f" ({loc})" if loc else ""
        items.append((f"🆕 {dn}{loc_str} 신규 입고, {_man(cur_price)} · {md}", _retail_url))
        seen_new += 1
        if seen_new >= 3:
            break
    seen_drops = 0
    for nm, prev_price, cur_price, delta, pct, _, loc, _, cf in r_drops:
        if cf or prev_price is None or cur_price is None:
            continue
        dn = _disp_name(nm)
        loc_str = f" ({loc})" if loc else ""
        items.append((
            f"📉 {dn}{loc_str} 가격 하락, "
            f"{_man(prev_price)}→{_man(cur_price)} ({abs(pct):.0f}% 하락) · {md}",
            _retail_url))
        seen_drops += 1
        if seen_drops >= 3:
            break

    # 🚚 트레이더스 OCR 최신 수집 이벤트 — 가끔 수집되므로 캐러셀에 뉴스로 노출.
    tr_total, tr_new, tr_drops, tr_date = _traders_ocr_signals()
    if tr_total:
        tr_md = _md(tr_date) if tr_date else md
        # trprice 전용 글 있으면 링크(없으면 _retail_url 폴백)
        import glob as _glob2
        _tr_posts = sorted(_glob2.glob(
            os.path.join(ROOT, "blog-md", "_posts", "*-youtube-traders-prices.md")))
        if _tr_posts:
            _b = os.path.basename(_tr_posts[-1])[:-3]
            _yy, _mm, _dd, *_sl = _b.split("-")
            _tr_url = f"/{_yy}/{_mm}/{_dd}/{'-'.join(_sl)}/"
        else:
            _tr_url = _retail_url
        # 수집 총수 알림 — 날짜가 오늘이 아닐 수도 있으므로 날짜 명시
        items.append((
            f"🚚 트레이더스 현장가 {tr_total}개 업데이트 · {tr_md}", _tr_url))
        # 신규 입고 TOP 2
        for tr_nm, tr_price, tr_store in tr_new[:2]:
            dn = _disp_name(tr_nm)
            store_str = f" ({tr_store})" if tr_store else ""
            items.append((
                f"🆕 {dn}{store_str} 신규 입고, {_man(tr_price)} · {tr_md}", _tr_url))
        # 가격 하락 TOP 2
        for tr_nm, tr_old, tr_cur, tr_pct in tr_drops[:2]:
            dn = _disp_name(tr_nm)
            items.append((
                f"📉 {dn} 트레이더스 인하, "
                f"{_man(tr_old)}→{_man(tr_cur)} ({tr_pct:.0f}% 할인) · {tr_md}", _tr_url))

    return items


def _carousel_front_matter(date_str):
    """compare 글 front matter 의 carousel YAML 블록 문자열(없으면 빈 문자열).

    각 아이템 = {text: "...", url: "/..."} dict — 홈 캐러셀이 per-item 링크 렌더.
    """
    items = build_carousel_items(date_str)
    if not items:
        return ""
    out = [f'carousel_date: "{date_str}"', "carousel:"]
    for text, url in items:
        esc_t = text.replace("\\", "\\\\").replace('"', '\\"')
        esc_u = url.replace("\\", "\\\\").replace('"', '\\"')
        out.append(f'  - text: "{esc_t}"')
        out.append(f'    url: "{esc_u}"')
    return "\n".join(out) + "\n"


def build_blog_md(both, date_str, blog_dir=None, crawl_stats=None):
    """면세 가격 비교 주간 로그 MD 생성.

    Args:
        both: 비교 행 목록(from build_comparison)
        date_str: 기준 날짜 YYYY-MM-DD
        blog_dir: 출력 디렉터리(기본=blog-md/_posts)
        crawl_stats: 수집 통계 dict — {
            '신라': {'count': N, 'date': 'YYYY-MM-DD'},
            '롯데': {'count': N, 'date': 'YYYY-MM-DD'},
            '신세계': {'count': N, 'date': 'YYYY-MM-DD'},
        }
    """
    EXCH = 1545.3
    GAP = 10.0
    # 보드(CMPA-666 2026-06-28): 가격대 구간 세분화.
    # 보드 2차(2026-06-28): 마지막 구간 30~50만원으로 상한 두고, 50만원 초과는 버림(브래킷 없음).
    BRACKETS = [
        ('10만원 이하',   0,       100000),
        ('10~15만원',  100000,   150000),
        ('15~20만원',  150000,   200000),
        ('20~30만원',  200000,   300000),
        ('30~50만원',  300000,   500000),
    ]
    SHOP = {'신라': '신라면세', '롯데': '롯데면세', '신세계': '신세계면세'}
    story_idx = load_story_lookup()

    from collections import defaultdict
    # 보드(CMPA-666 2026-06-28): 절약(100ml당) = (국내 소매가 − 면세 최저가)를 100ml당 KRW로 환산.
    # 양수면 면세가 그만큼 싸다. 이 값 내림차순(국내 소매가 없는 행은 맨 아래)으로 가격대별 Top 10.
    dom_dates = sorted({r.get('국내수집일', '') for r in both if r.get('국내수집일')})
    bracket_rows = defaultdict(list)
    for r in both:
        prices = {s: r.get(f'{s}_USD') for s in ('신라', '롯데', '신세계')}
        vals = {s: v for s, v in prices.items() if v not in ('', None)}
        if not vals:
            continue
        min_usd = min(vals.values())
        min_krw = min_usd * EXCH
        cheapest = r.get('최저가_면세점', '')
        dp100 = r.get('국내최저_₩100ml')  # 국내 100ml당(소매 floor, KRW)
        dp100 = float(dp100) if dp100 not in ('', None) else None
        # 최저가 면세점의 용량(국내 소매가를 같은 용량으로 환산하기 위함)
        min_shop = min(vals, key=vals.get)
        duty_vol = None
        try:
            duty_vol = int(r.get(f'{min_shop}용량ml') or 0) or None
        except (ValueError, TypeError):
            duty_vol = None
        # 절약(100ml당) = 국내 소매가(100ml당) − 면세 최저가(100ml당). 양수=면세가 더 쌈.
        save_p100 = None
        if dp100 is not None and duty_vol:
            duty_p100_krw = min_usd / duty_vol * 100.0 * EXCH
            save_p100 = int(round(dp100 - duty_p100_krw))
        for bname, lo, hi in BRACKETS:
            if lo <= min_krw < hi:
                bracket_rows[bname].append({
                    'name': r['제품명'], 'vol': r.get('용량ml', ''),
                    'cheapest': cheapest,
                    'dom_p100': dp100,
                    'duty_vol': duty_vol,
                    'min_usd': min_usd, 'min_krw': int(min_krw),
                    '신라': prices['신라'], '롯데': prices['롯데'], '신세계': prices['신세계'],
                    'save_p100': save_p100,
                    'dom_src': r.get('국내최저_소스', ''),
                    'dom_tpid': r.get('국내최저_tpid', ''),
                    'dom_krw': r.get('국내최저_₩', ''),
                    'dom_vol': r.get('국내최저_용량ml', ''),
                })
                break

    def fu(v): return f'${v:,.2f}' if v not in ('', None) else '—'
    def bold(shop, item):
        v = item[shop]
        s = fu(v)
        return f'<strong>{s}</strong>' if shop == item['cheapest'] else s

    yr_mo = date_str[:7].replace('-', '년 ') + '월'
    dom_dates_str = ' · '.join(dom_dates) or '-'
    # 캐러셀 아이템(면세 vs 국내 gap 하이라이트) — 홈 _HOME_CAROUSEL 이 이 front matter 를 읽어 회전 노출.
    car_fm = _carousel_front_matter(date_str)
    lines = [f'''---
layout: post
title: "면세 위스키 가격 비교 주간 로그 (신라·롯데·신세계, {yr_mo})"
date: {date_str}
categories: [wprice]
kind: wprice
data_date: "{date_str}"
robots: "index,follow"
{car_fm}---

이 글은 신라·롯데·신세계 세 면세점의 위스키 가격을 **주기적으로 수집해 아래에 쌓는 로그**입니다. 가격대별 Top 10은 매 수집 시 갱신되고, 아래 수집 로그는 최신이 맨 위입니다.

> **면세 조건**: 1인 출국 시 2병/2L/$400 한도. 가격은 USD 표시가 기준.
> **100ml당 USD**로 비교해 같은 기준으로 면세점을 비교합니다.

---

## 💰 가격대별 Top 10 — 면세 vs 국내 100ml당 절약 큰 순

> **최저가(면세점)** = 세 면세점 중 가장 싼 곳·가격 (볼드).
> **절약(100ml당)** = (국내 소매가 − 면세 최저가)를 **100ml당 금액(원)**으로 환산한 값 — **양수면 면세가 그만큼 더 싸다.** 이 값이 큰 순서로 가격대별 Top 10. 국내 소매가가 없으면 `—`.
> **국내 소매가** = 데일리샷·트레이더스·코스트코 중 **100ml당 최저가의 실제 판매가**(괄호=출처). 국내 판매 용량이 면세 용량과 다르면 `·750ml`처럼 용량을 함께 표기합니다(가짜 환산가 대신 실가격). 절약(100ml당)은 용량을 100ml로 맞춰 공정 비교합니다. 없으면 `—`. 출처가 **(데일리샷)**이면 눌러서 해당 상품 페이지를 새 창으로 볼 수 있습니다.
> 💡 **제품명을 누르면** 스토리·맛 노트가 있는 경우 펼쳐집니다.
> 국내 소매가 수집일: {dom_dates_str}. 환율 1 USD = 1,545원 단일 적용.

''']

    def save_cell(v):
        if v is None:
            return '—'
        if v >= 0:
            return f'+₩{v:,}'
        return f'−₩{abs(v):,}'

    def retail_cell_md(item):
        """국내 소매가 셀 = 실제 판매 가격·용량(면세와 용량 다르면 ·용량 표기, 공정비교는 절약 100ml당)."""
        dom_krw = item.get('dom_krw')
        if not dom_krw:
            return '—'
        try:
            won = int(round(float(dom_krw)))
        except (ValueError, TypeError):
            return '—'
        src_html = dom_src_html(item['dom_src'], item['name'], item.get('dom_tpid'))
        # 면세 최저가 용량과 국내 판매 용량이 다르면 용량을 병기(가짜 환산가 표시 방지).
        dvol_dom, duty_vol = item.get('dom_vol'), item.get('duty_vol')
        vol_note = ''
        try:
            if dvol_dom and duty_vol and int(dvol_dom) != int(duty_vol):
                vol_note = f'·{int(dvol_dom)}ml'
        except (ValueError, TypeError):
            vol_note = ''
        return f'₩{won:,}{vol_note} {src_html}'

    def product_cell_md(item):
        return story_cell_html(item['name'], story_idx)

    for bname, lo, hi in BRACKETS:
        # 절약(100ml당) 내림차순(국내 소매가 없는 행은 맨 아래), 가격대별 Top 10
        items = sorted(
            bracket_rows[bname],
            key=lambda x: (x['save_p100'] is not None,
                           x['save_p100'] if x['save_p100'] is not None else 0),
            reverse=True)[:10]
        if not items:
            continue
        # 제품 셀에 스토리(<details>)가 들어가므로 표는 raw HTML로 생성한다.
        # (kramdown 파이프 표 셀은 블록 HTML <details>/<summary>를 이스케이프해 깨짐 —
        #  raw <table> 블록은 그대로 통과되어 셀 안에서 정상 렌더된다. .post table CSS 동일 적용.)
        lines.append(f'### {bname}\n')
        h = ('<table><thead><tr>'
             '<th>제품</th><th>최저가 (면세점)</th><th>절약(100ml당)</th>'
             '<th>국내 소매가</th><th>신라</th><th>롯데</th><th>신세계</th><th>용량</th>'
             '</tr></thead><tbody>')
        rows_html = []
        for item in items:
            cheapest_label = SHOP.get(item['cheapest'], item['cheapest'])
            min_str = (f"<strong>{fu(item['min_usd'])} / ₩{item['min_krw']:,}</strong> "
                       f"({cheapest_label})")
            rows_html.append(
                '<tr>'
                f'<td>{product_cell_md(item)}</td>'
                f'<td>{min_str}</td>'
                f'<td>{save_cell(item["save_p100"])}</td>'
                f'<td>{retail_cell_md(item)}</td>'
                f'<td>{bold("신라", item)}</td>'
                f'<td>{bold("롯데", item)}</td>'
                f'<td>{bold("신세계", item)}</td>'
                f'<td>{item["vol"]}ml</td>'
                '</tr>')
        lines.append(h + ''.join(rows_html) + '</tbody></table>')
        lines.append('')

    # 수집 로그 섹션
    lines.append('---\n\n## 📋 수집 로그 (최신이 맨 위)\n\n---\n')

    stats = crawl_stats or {
        '신라': {'count': '?', 'date': SHILLA_DATE},
        '롯데': {'count': '?', 'date': LOTTE_DATE},
        '신세계': {'count': '?', 'date': SSG_DATE},
    }
    total_matched = len(both)
    all3 = sum(1 for r in both
               if all(r.get(f'{s}_USD') not in ('', None) for s in ('신라', '롯데', '신세계')))

    shilla_d = stats['신라']['date']
    lotte_d = stats['롯데']['date']
    ssg_d = stats['신세계']['date']
    shilla_n = stats['신라']['count']
    lotte_n = stats['롯데']['count']
    ssg_n = stats['신세계']['count']

    lines.append(f'''
## 📅 {date_str} — 수집 ({shilla_d} 신라 · {lotte_d} 롯데 · {ssg_d} 신세계)

### 수집 결과

| 면세점 | 수집 종수 | 수집일 |
|---|---|---|
| 신라면세 | {shilla_n}종 | {shilla_d} |
| 롯데면세 | {lotte_n}종 | {lotte_d} |
| 신세계면세 | {ssg_n}종 | {ssg_d} |

- 세 면세점 동시 비교 가능: **{total_matched}종** (3곳 모두 있는 제품: {all3}종)
- 가격차 10% 초과 제품: 위 가격대별 Top 10 표 참고

*by Dram · CaskCode*
''')

    md = '\n'.join(lines)
    if blog_dir is None:
        blog_dir = os.path.join(ROOT, 'blog-md', '_posts')
    os.makedirs(blog_dir, exist_ok=True)
    out = os.path.join(blog_dir, f'{date_str}-dutyfree-whisky-compare.md')
    with open(out, 'w', encoding='utf-8') as f:
        f.write(md)
    return out


def build_mart_blog_md(both, date_str, blog_dir=None, crawl_stats=None):
    """'마트에서 구매할 때' 페이지 — 면세점보다 싸거나 비슷하게 국내(마트·온라인)에서 살 수 있는
    위스키 (CMPA-666 보드 확장). dutyfree-compare 와 같은 데이터(build_rows)·셀 렌더를 재사용.

    포함 기준: 국내최저(데일리샷+트레이더스+코스트코) 100ml당가 ≤ 면세 최저 100ml당가 × 1.05
    (국내가 더 싸거나 ~5% 이내로 비슷). 오매칭 백스톱(양방향 2.5배)·번들/멀티팩 노이즈 제외.
    국내가 면세보다 싼 정도(100ml당) 내림차순.
    """
    EXCH = 1545.3
    SIM = 1.05            # 국내 ≤ 면세×1.05 (싸거나 비슷)
    story_idx = load_story_lookup()
    SHOPN = {'신라': '신라면세', '롯데': '롯데면세', '신세계': '신세계면세'}

    rows = []
    for r in both:
        nm = r['제품명']
        if _MULTIPACK_PAT.search(nm.lower()) or is_bundle_noise(nm):
            continue
        prices = {s: r.get(f'{s}_USD') for s in ('신라', '롯데', '신세계')}
        vals = {s: v for s, v in prices.items() if v not in ('', None)}
        if not vals:
            continue
        min_usd = min(vals.values())
        min_shop = min(vals, key=vals.get)
        try:
            dvol = int(r.get(f'{min_shop}용량ml') or 0) or None
        except (ValueError, TypeError):
            dvol = None
        dp100 = r.get('국내최저_₩100ml')
        if dp100 in ('', None) or not dvol:
            continue
        dp100 = float(dp100)
        duty100 = min_usd / dvol * 100.0 * EXCH
        if dp100 > duty100 * SIM:        # 국내가 면세보다 5% 넘게 비싸면 제외
            continue
        rows.append({
            'name': nm, 'vol': r.get('용량ml', ''),
            'min_usd': min_usd, 'min_krw': int(round(min_usd * EXCH)),
            'min_shop': SHOPN.get(min_shop, min_shop),
            'duty_vol': dvol,
            'dom_krw': r.get('국내최저_₩', ''), 'dom_src': r.get('국내최저_소스', ''),
            'dom_tpid': r.get('국내최저_tpid', ''), 'dom_vol': r.get('국내최저_용량ml', ''),
            'adv100': int(round(duty100 - dp100)),     # 양수=국내가 그만큼 싸다(100ml당)
        })
    # 같은 위스키의 용량 변형(예: 발렌타인 21년 500/700ml)은 한 줄로 — 100ml당 이득이 가장 큰 행만.
    _volkey = re.compile(r"\d+(?:ml|l)$")
    best = {}
    for it in rows:
        k = _volkey.sub("", norm(it['name']))
        if k not in best or it['adv100'] > best[k]['adv100']:
            best[k] = it
    rows = list(best.values())
    rows.sort(key=lambda x: -x['adv100'])

    n_mart = sum(1 for x in rows if x['dom_src'] in ('트레이더스', '코스트코'))

    def adv_cell(v):
        if v >= 1000:
            return f'<strong>국내 +₩{v:,}</strong>'
        if v <= -1000:
            return f'면세 +₩{abs(v):,}'
        return '≈ 비슷'

    def retail(item):
        try:
            won = int(round(float(item['dom_krw'])))
        except (ValueError, TypeError):
            return '—'
        sh = dom_src_html(item['dom_src'], item['name'], item.get('dom_tpid'))
        vn = ''
        try:
            if item['dom_vol'] and item['duty_vol'] and int(item['dom_vol']) != int(item['duty_vol']):
                vn = f"·{int(item['dom_vol'])}ml"
        except (ValueError, TypeError):
            vn = ''
        return f'₩{won:,}{vn} {sh}'

    yr_mo = date_str[:7].replace('-', '년 ') + '월'
    dom_dates = sorted({r.get('국내수집일', '') for r in both if r.get('국내수집일')})
    dom_dates_str = ' · '.join(dom_dates) or '-'
    lines = [f'''---
layout: post
title: "마트에서 사는 게 이득인 위스키 — 면세점보다 싸거나 비슷한 ({yr_mo})"
date: {date_str}
categories: [wprice]
kind: wprice
data_date: "{date_str}"
robots: "index,follow"
---

면세점이 항상 가장 싸지는 않습니다. **트레이더스·코스트코 같은 마트나 국내 온라인(데일리샷)에서 면세가보다 싸거나 비슷하게** 살 수 있는 위스키를 모았습니다. 면세 한도(1인 2병·$400)를 다른 술에 쓰고 싶을 때, 이 목록은 **굳이 면세점을 기다리지 않아도 되는** 위스키입니다.

> **상세 표기** — `국내`: 데일리샷·트레이더스·코스트코 중 100ml당 최저가(괄호=출처, 🛒=마트, 용량이 다르면 `·750ml` 병기). `면세`: 신라·롯데·신세계 중 최저. 마지막 줄 = 100ml 기준 국내·면세 비교(양수=국내가 이득, `≈ 비슷`=차이 작음). **(데일리샷)** 은 눌러서 상품 페이지로.
> 💡 위스키 이름을 누르면 스토리·맛 노트가 있는 경우 펼쳐집니다.
> 국내 수집일: {dom_dates_str}. 환율 1 USD = 1,545원 단일 적용. 면세가는 출국·면세한도 조건의 '수집일 기준'값입니다.

---

## 🛒 면세점보다 싸거나 비슷한 위스키 ({len(rows)}종 · 국내가 싼 순)
''']
    if not rows:
        lines.append('\n*이번 수집에서는 해당 제품이 없습니다.*\n')
    else:
        # 2-column 모바일 우선 레이아웃 (CMPA-769): (위스키, 상세)
        # 상세 = 국내최저 / 면세최저 / 100ml당 비교 + 용량 — <br>로 줄바꿈
        h = ('<table><thead><tr>'
             '<th>위스키</th><th>상세</th>'
             '</tr></thead><tbody>')
        body = []
        for it in rows:
            dom_line = f"국내 {retail(it)}"
            duty_line = (f"면세 <strong>${it['min_usd']:,.2f} / "
                         f"₩{it['min_krw']:,}</strong> ({it['min_shop']})")
            adv_line = f"{adv_cell(it['adv100'])} · {it['vol']}ml"
            detail = f"{dom_line}<br>{duty_line}<br>{adv_line}"
            body.append(
                '<tr>'
                f"<td>{story_cell_html(it['name'], story_idx)}</td>"
                f"<td style='font-size:.93em'>{detail}</td>"
                '</tr>')
        lines.append(h + ''.join(body) + '</tbody></table>')
        lines.append('')

    stats = crawl_stats or {
        '신라': {'count': '?', 'date': SHILLA_DATE},
        '롯데': {'count': '?', 'date': LOTTE_DATE},
        '신세계': {'count': '?', 'date': SSG_DATE},
    }
    lines.append(f'''
---

## 📋 수집 기준

| 소스 | 수집일 |
|---|---|
| 국내 최저(데일리샷·트레이더스·코스트코) | {dom_dates_str} |
| 신라면세 | {stats['신라']['date']} |
| 롯데면세 | {stats['롯데']['date']} |
| 신세계면세 | {stats['신세계']['date']} |

- 면세점보다 싸거나 비슷한 국내가 위스키 **{len(rows)}종** (그중 마트=트레이더스·코스트코 최저 **{n_mart}종**).
- 같은 제품이 아닐 가능성이 높은 오매칭(국내가가 면세의 2.5배↑ 또는 1/2.5↓)·번들/잔세트는 제외했습니다.

*by Dram · CaskCode*
''')

    md = '\n'.join(lines)
    if blog_dir is None:
        blog_dir = os.path.join(ROOT, 'blog-md', '_posts')
    os.makedirs(blog_dir, exist_ok=True)
    out = os.path.join(blog_dir, f'{date_str}-mart-cheaper-whisky.md')
    with open(out, 'w', encoding='utf-8') as f:
        f.write(md)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--date", default=LOTTE_DATE, help="산출 파일명 날짜(기본=롯데 수집일)")
    ap.add_argument("--blog", action="store_true", help="블로그 MD 초안 생성 (blog-md/_posts/)")
    args = ap.parse_args()

    # staleness 가드(CMPA-672): 최신 스냅샷이 오래되면 경고만(빌드는 최신 가용본으로 계속).
    print(f"스냅샷: 신라 {SHILLA_DATE} · 롯데 {LOTTE_DATE} · 신세계 {SSG_DATE} · "
          f"데일리샷보강 {ONLINE_DATE or '없음'}")
    warn_all_staleness()

    both, s_only, l_only, g_only, n_s, n_l, n_g, dom_meta = build_rows()

    outdir = os.path.join(ROOT, "reports", "dutyfree-compare")
    os.makedirs(outdir, exist_ok=True)
    base = f"면세점_100ml_가격비교_{args.date}"
    xlsx_path = os.path.join(outdir, base + ".xlsx")
    csv_path = os.path.join(outdir, base + ".csv")

    cmp_cols = write_xlsx(xlsx_path, both, s_only, l_only, g_only, n_s, n_l, n_g, dom_meta)
    write_csv(csv_path, both, cmp_cols)

    all3 = sum(1 for r in both if r["신라_USD"] != "" and r["롯데_USD"] != "" and r["신세계_USD"] != "")
    sl_pair = sum(1 for r in both if r["신세계_USD"] == "" and r["신라_USD"] != "" and r["롯데_USD"] != "")
    sg_pair = sum(1 for r in both if r["롯데_USD"] == "" and r["신라_USD"] != "" and r["신세계_USD"] != "")
    lg_pair = sum(1 for r in both if r["신라_USD"] == "" and r["롯데_USD"] != "" and r["신세계_USD"] != "")

    print(f"신라 매칭후보 {n_s}행 · 롯데 매칭후보 {n_l}행 · 신세계 매칭후보 {n_g}행")
    print(f"비교(교집합) {len(both)}종 (2개 이상 면세점 매칭)")
    print(f"  → 3개 모두: {all3}  신라+롯데만: {sl_pair}  신라+신세계만: {sg_pair}  롯데+신세계만: {lg_pair}")
    print(f"신라단독 {len(s_only)} · 롯데단독 {len(l_only)} · 신세계단독 {len(g_only)}")
    dom_n = sum(1 for r in both if r.get("국내최저_₩100ml") not in ("", None))
    duty_win = sum(1 for r in both if r.get("최저처(면세vs국내)") == "면세")
    dom_win = sum(1 for r in both if r.get("최저처(면세vs국내)") == "국내")
    print(f"국내최저가(데일리샷+트레이더스+코스트코) 매칭 {dom_n}종 "
          f"— 면세가 더 쌈 {duty_win} / 국내가 더 쌈 {dom_win} (환율 {dom_meta['fx']:,.0f})")
    if args.blog:
        import csv as _csv
        def _raw_count(path):
            try:
                with open(path, encoding='utf-8-sig') as f:
                    return sum(1 for _ in _csv.DictReader(f))
            except Exception:
                return '?'
        crawl_stats = {
            '신라': {'count': _raw_count(SHILLA_CSV), 'date': SHILLA_DATE},
            '롯데': {'count': _raw_count(LOTTE_CSV), 'date': LOTTE_DATE},
            '신세계': {'count': _raw_count(SSG_CSV), 'date': SSG_DATE},
        }
        blog_path = build_blog_md(both, args.date, crawl_stats=crawl_stats)
        print(f"blog -> {blog_path}")
        mart_path = build_mart_blog_md(both, args.date, crawl_stats=crawl_stats)
        print(f"mart -> {mart_path}")

    print(f"xlsx -> {xlsx_path}")
    print(f"csv  -> {csv_path}")
    print("\n=== 3-면세점 가격차 TOP 12 (100ml당 USD 스프레드) ===")
    for r in both[:12]:
        s_str = f"신라 ${r['신라_USD100ml']:>6.2f}" if r["신라_USD100ml"] != "" else "신라 N/A    "
        l_str = f"롯데 ${r['롯데_USD100ml']:>6.2f}" if r["롯데_USD100ml"] != "" else "롯데 N/A    "
        g_str = f"신세계 ${r['신세계_USD100ml']:>6.2f}" if r["신세계_USD100ml"] != "" else "신세계 N/A    "
        print(f"{r['가격차_%']:5.1f}%  {r['제품명']:30} {s_str} | {l_str} | {g_str} → 최저={r['최저가_면세점']}")


if __name__ == "__main__":
    main()
